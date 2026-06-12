import io
import zipfile
import csv
from datetime import datetime

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from database import supabase

router = APIRouter()

_FILL = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
_FONT = Font(color="FFFFFF", bold=True)

_TAX_TYPE_KO = {
    'pension_savings':    '연금저축',
    'retirement_pension': '퇴직연금(IRP)',
    'isa':               'ISA',
    'regular':           '일반',
}


def _header_style(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = _FILL
        cell.font = _FONT


def _auto_width(ws):
    for col in ws.columns:
        width = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 50)


def _fetch():
    assets = (
        supabase.table("assets").select("*")
        .order("account_name").order("asset_name").execute().data or []
    )
    incomes = (
        supabase.table("income_log").select("*")
        .order("income_date", desc=True).execute().data or []
    )
    withdrawals = (
        supabase.table("withdrawals").select("*")
        .order("withdrawal_date", desc=True).execute().data or []
    )
    expenses = (
        supabase.table("expenses").select("*")
        .order("expense_date", desc=True).execute().data or []
    )
    cfg_rows = (
        supabase.table("user_config").select("value")
        .eq("key", "config").execute().data or [{}]
    )
    cfg = cfg_rows[0].get("value", {}) if cfg_rows else {}
    return assets, incomes, withdrawals, expenses, cfg


def _cfg_items(cfg: dict):
    user = cfg.get("user", {})
    portfolio = cfg.get("portfolio", {})
    birth_year = user.get("birth_year") or 0
    age = (datetime.now().year - int(birth_year)) if birth_year else ""
    return [
        ("현재 나이",      age),
        ("은퇴 나이",      user.get("retirement_age", "")),
        ("월 생활비 (원)", user.get("monthly_expense", "")),
        ("Cash 목표%",    round(float(portfolio.get("target_cash",   0)) * 100, 1)),
        ("Bond 목표%",    round(float(portfolio.get("target_bond",   0)) * 100, 1)),
        ("Equity 목표%",  round(float(portfolio.get("target_equity", 0)) * 100, 1)),
        ("Income 목표%",  round(float(portfolio.get("target_income", 0)) * 100, 1)),
        ("알림 이메일",    cfg.get("alert", {}).get("email", "")),
    ]


def _sheet_summary(ws, assets: list, cfg: dict, now_str: str):
    portfolio = cfg.get("portfolio", {})
    active = [a for a in assets if a.get("is_active")]
    total = sum(float(a.get("current_value") or 0) for a in active)

    buckets: dict[str, float] = {"cash": 0.0, "bond": 0.0, "tdf": 0.0, "equity": 0.0, "income": 0.0}
    for a in active:
        t = (a.get("asset_type") or "").lower()
        v = float(a.get("current_value") or 0)
        if t in buckets:
            buckets[t] += v
        elif t == "fund":
            buckets["bond"] += v

    targets = {
        "cash":   float(portfolio.get("target_cash",   0)),
        "bond":   float(portfolio.get("target_bond",   0)),
        "tdf":    0.0,
        "equity": float(portfolio.get("target_equity", 0)),
        "income": float(portfolio.get("target_income", 0)),
    }

    ws.append(["내보내기 일시", now_str])
    ws.append(["총 자산 (원)", total])
    ws.append([])
    ws.append(["버킷", "현재 비중 (%)", "목표 비중 (%)"])
    _header_style(ws, row=4)

    for key, label in [
        ("cash",   "Cash (현금성)"),
        ("bond",   "Bond (채권)"),
        ("tdf",    "TDF"),
        ("equity", "Equity (주식형)"),
        ("income", "Income (리츠/인컴)"),
    ]:
        cur = round(buckets[key] / total * 100, 1) if total > 0 else 0
        tgt = round(targets[key] * 100, 1) if targets[key] else "-"
        ws.append([label, cur, tgt])

    _auto_width(ws)


def _sheet_assets(ws, assets: list):
    ws.append(["계좌명", "자산명", "자산유형", "세제분류", "수량", "현재가", "평가액", "매입일", "만기일", "활성여부"])
    _header_style(ws)
    for a in assets:
        ws.append([
            a.get("account_name", ""),
            a.get("asset_name", ""),
            a.get("asset_type", ""),
            _TAX_TYPE_KO.get(a.get("tax_account_type") or "", "미분류"),
            float(a.get("quantity") or 0),
            float(a.get("unit_price") or 0),
            float(a.get("current_value") or 0),
            a.get("purchase_date") or "",
            a.get("maturity_date") or "",
            "활성" if a.get("is_active") else "비활성",
        ])
    _auto_width(ws)


def _sheet_income(ws, incomes: list):
    ws.append(["날짜", "유형", "금액", "자산명", "메모"])
    _header_style(ws)
    for inc in incomes:
        ws.append([
            inc.get("income_date", ""),
            inc.get("income_type", ""),
            float(inc.get("amount") or 0),
            inc.get("asset_name", ""),
            inc.get("note") or "",
        ])
    _auto_width(ws)


_CATEGORY_KO = {
    'living':  '생활비',
    'housing': '주거·관리',
    'medical': '의료·건강',
    'family':  '경조사·가족',
    'leisure': '여행·여가',
    'other':   '기타',
}


def _sheet_expenses(ws, expenses: list):
    ws.append(["날짜", "카테고리", "금액", "메모"])
    _header_style(ws)
    for e in expenses:
        ws.append([
            e.get("expense_date", ""),
            _CATEGORY_KO.get(e.get("category") or "other", "기타"),
            float(e.get("amount") or 0),
            e.get("memo") or "",
        ])
    _auto_width(ws)


def _sheet_withdrawals(ws, withdrawals: list):
    ws.append(["날짜", "계좌명", "세제분류", "금액", "메모"])
    _header_style(ws)
    for w in withdrawals:
        ws.append([
            w.get("withdrawal_date", ""),
            w.get("account_name", ""),
            _TAX_TYPE_KO.get(w.get("tax_account_type") or "", "미분류"),
            float(w.get("amount") or 0),
            w.get("memo") or "",
        ])
    _auto_width(ws)


def _sheet_config(ws, cfg: dict):
    ws.append(["항목", "값"])
    _header_style(ws)
    for row in _cfg_items(cfg):
        ws.append(list(row))
    _auto_width(ws)


@router.get("/xlsx")
def export_xlsx():
    today = datetime.now().strftime("%Y%m%d")
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    assets, incomes, withdrawals, expenses, cfg = _fetch()

    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "📊 요약"
    _sheet_summary(ws0, assets, cfg, now_str)

    _sheet_assets(wb.create_sheet("자산 목록"), assets)
    _sheet_income(wb.create_sheet("수입 기록"), incomes)
    _sheet_expenses(wb.create_sheet("지출 기록"), expenses)
    _sheet_withdrawals(wb.create_sheet("인출 기록"), withdrawals)
    _sheet_config(wb.create_sheet("설정"), cfg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"retirement_backup_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/csv")
def export_csv():
    today = datetime.now().strftime("%Y%m%d")

    assets, incomes, withdrawals, expenses, cfg = _fetch()

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # assets.csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["계좌명", "자산명", "자산유형", "세제분류", "수량", "현재가", "평가액", "매입일", "만기일", "활성여부"])
        for a in assets:
            w.writerow([
                a.get("account_name", ""),
                a.get("asset_name", ""),
                a.get("asset_type", ""),
                _TAX_TYPE_KO.get(a.get("tax_account_type") or "", "미분류"),
                a.get("quantity", 0) or 0,
                a.get("unit_price", 0) or 0,
                a.get("current_value", 0) or 0,
                a.get("purchase_date") or "",
                a.get("maturity_date") or "",
                "활성" if a.get("is_active") else "비활성",
            ])
        zf.writestr("assets.csv", buf.getvalue().encode("utf-8-sig"))

        # income.csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["날짜", "유형", "금액", "자산명", "메모"])
        for inc in incomes:
            w.writerow([
                inc.get("income_date", ""),
                inc.get("income_type", ""),
                inc.get("amount", 0) or 0,
                inc.get("asset_name", ""),
                inc.get("note") or "",
            ])
        zf.writestr("income.csv", buf.getvalue().encode("utf-8-sig"))

        # expenses.csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["날짜", "카테고리", "금액", "메모"])
        for e in expenses:
            w.writerow([
                e.get("expense_date", ""),
                _CATEGORY_KO.get(e.get("category") or "other", "기타"),
                e.get("amount", 0) or 0,
                e.get("memo") or "",
            ])
        zf.writestr("expenses.csv", buf.getvalue().encode("utf-8-sig"))

        # withdrawals.csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["날짜", "계좌명", "세제분류", "금액", "메모"])
        for wd in withdrawals:
            w.writerow([
                wd.get("withdrawal_date", ""),
                wd.get("account_name", ""),
                _TAX_TYPE_KO.get(wd.get("tax_account_type") or "", "미분류"),
                wd.get("amount", 0) or 0,
                wd.get("memo") or "",
            ])
        zf.writestr("withdrawals.csv", buf.getvalue().encode("utf-8-sig"))

        # config.csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["항목", "값"])
        for row in _cfg_items(cfg):
            w.writerow(list(row))
        zf.writestr("config.csv", buf.getvalue().encode("utf-8-sig"))

    zip_buf.seek(0)
    filename = f"retirement_backup_{today}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
